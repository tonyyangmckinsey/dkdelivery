import pandas as pd
import os
import glob
import re
import time
from openpyxl import load_workbook
from metric_avg_days_in_step import calculate_avg_days_in_step
from metric_avg_orders_per_day import calculate_avg_orders_per_day
from metric_avg_mrc_per_day import calculate_avg_mrc_per_day
from metric_orders_in_out import calculate_orders_in_out
from individual_avg_orders_per_day import calculate_avg_orders_per_day_per_dpm
from individual_avg_mrc_per_day import calculate_avg_mrc_per_day_per_dpm
from individual_avg_days_in_step import calculate_avg_days_in_step_per_dpm
from individual_orders_in_out import calculate_orders_in_out_per_dpm

def normalize_keys(df):
    df["ServiceID_Crid"] = df["ServiceID_Crid"].astype(str).str.strip()
    df["SalesProjectID_ContractNumber"] = df["SalesProjectID_ContractNumber"].astype(str).str.strip()
    return df


def extract_date_columns(df):
    return [col for col in df.columns if re.match(r"^\d{8}$", str(col)) and df[col].notna().any()]


def analyze_order_development(file_path):

    print("HEII")
    df = pd.read_excel(file_path, engine="openpyxl")
    df = normalize_keys(df)
    date_columns = extract_date_columns(df)
    print("Her er datecolumns", date_columns)

    steps = [
        "1. Order entry and validation",
        "1.9 PMO delivery planning",
        "2. Delivery planning",
        "3. Installation preparation and digging order",
        "4. Final design",
        "5. Configuration",
        "6. Cabling splicing and installation",
        "7. Service activation",
        "8. Ready for service",
        "9. Vendor invoice",
        "Delivered",
        "Cancelled/terminated/On hold"
    ]

    tracked_steps = [
        "1.9 PMO delivery planning",
        "2. Delivery planning",
        "3. Installation preparation and digging order",
        "4. Final design",
        "5. Configuration",
        "6. Cabling splicing and installation",
        "7. Service activation",
        "8. Ready for service"
    ]

    avg_days_df = calculate_avg_days_in_step(df, date_columns, steps)
    avg_orders_per_day_df = calculate_avg_orders_per_day(df, date_columns, steps)
    avg_mrc_per_day_df = calculate_avg_mrc_per_day(df, date_columns, steps)
    summary_df = avg_days_df.merge(avg_orders_per_day_df, on="Step").merge(avg_mrc_per_day_df, on="Step")

    T1 = input("Enter T1 date for order in & out analysis (format: YYYYMMDD) (Earliest 20250519): ")
    T2 = input("Enter T2 date for order in & out analysis (format: YYYYMMDD): ")
    in_out_df = calculate_orders_in_out(df, date_columns, steps, T1, T2)

    avg_orders_per_dpm_df = calculate_avg_orders_per_day_per_dpm(df, date_columns, steps)
    avg_mrc_per_dpm_df = calculate_avg_mrc_per_day_per_dpm(df, date_columns, steps)
    avg_days_per_dpm_df = calculate_avg_days_in_step_per_dpm(df, date_columns, steps)

    output_path = os.path.join(
        r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\FinalOutputFolder",
        os.path.basename(file_path).replace(".xlsx", "_step_metrics_summary.xlsx")
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Step Metrics Summary")
        in_out_df.to_excel(writer, index=False, sheet_name="Step Entry-Exit Summary")
        avg_orders_per_dpm_df.to_excel(writer, index=False, sheet_name="DPM Avg Orders per Day")
        avg_mrc_per_dpm_df.to_excel(writer, index=False, sheet_name="DPM Avg MRC per Day")
        avg_days_per_dpm_df.to_excel(writer, index=False, sheet_name="DPM Avg Days in Step")

        for step in tracked_steps:
            per_step_df = calculate_orders_in_out_per_dpm(df, date_columns, step, T1, T2)
            safe_sheet_name = step[:31]  # Excel max sheet name length is 31
            writer.book.create_sheet(safe_sheet_name)
            per_step_df.to_excel(writer, index=False, sheet_name=safe_sheet_name)

    wb = load_workbook(output_path)
    sheet1 = wb["Step Metrics Summary"]
    sheet2 = wb["Step Entry-Exit Summary"]
    sheet3 = wb["DPM Avg Orders per Day"]
    sheet4 = wb["DPM Avg MRC per Day"]
    sheet5 = wb["DPM Avg Days in Step"]

    start_date = "20250519"
    today_date = date_columns[-1]

    sheet1["A15"] = f"Data as of {start_date} - {today_date}"
    sheet2["A15"] = f"Interval between T1 and T2: {T1} - {T2}"
    sheet3["A28"] = f"Data as of {start_date} - {today_date}"
    sheet4["A28"] = f"Data as of {start_date} - {today_date}"
    sheet5["A28"] = f"Data as of {start_date} - {today_date}"

    wb.save(output_path)
    print(f"✅ Summary exported to: {output_path}")


def get_latest_order_output(folder):
    pattern = os.path.join(folder, "*_Order_Development_Output.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError("❌ No Order_Development_Output.xlsx files found.")
    latest_file = max(files, key=os.path.getmtime)
    return latest_file


def run_analysis():
    start_time = time.time()
    output_folder = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\AnalysisOutputFolder"
    file_path = get_latest_order_output(output_folder)
    analyze_order_development(file_path)
    end_time = time.time()
    print(f"⏱️ Execution time: {end_time - start_time:.2f} seconds")


if __name__ == "__main__":
    run_analysis()
