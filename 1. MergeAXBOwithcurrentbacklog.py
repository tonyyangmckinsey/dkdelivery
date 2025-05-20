import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
today_str = datetime.today().strftime("%Y%m%d")  # Format: YYYYMMDD


# === Load and clean main CSV ===
main_path = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\1. Input\axbo.csv"

df_raw = pd.read_csv(main_path, header=None, low_memory=False)
df_raw.columns = df_raw.iloc[0]
df = df_raw.drop(index=0).reset_index(drop=True)

# === Filter rows where "Committed RFS" >= 2023-01-01 ===
#df["Committed RFS"] = pd.to_datetime(df["Committed RFS"], errors='coerce')  # Convert to datetime
#df = df[df["Committed RFS"] >= pd.Timestamp("2023-01-01")]  # Keep only rows after 1 Jan 2023
#output_pathdf123 = r"C:\Users\tonyan\OneDrive - GlobalConnect A S\Documents\Output files\rawrawDF.xlsx"
#df.to_excel(output_pathdf123, index=False)

# Filter for relevant Project Stages
print("🧪 Unique Project Stages in raw file:", df["Project Stage"].unique())
df = df[df["Project Stage"].isin(["In Process", "On Hold", "Created"])]
output_pathdf = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\2. Output\rawDF.xlsx"
df.to_excel(output_pathdf, index=False)

# === Load Enterprise and SMB Excel files ===
enterprise_path = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\1. Input\Enterprisepublic.xlsx"
smb1_path = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\1. Input\smb1.xlsx"
smb2_path = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\1. Input\smb2.xlsx"
smb3_path = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\1. Input\smb3.xlsx"
#gcc_path = r"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python input\gcc.xlsx"

enterprise_df = pd.read_excel(enterprise_path, engine='openpyxl')
smb1_df = pd.read_excel(smb1_path, engine='openpyxl')
smb2_df = pd.read_excel(smb2_path, engine='openpyxl')
smb3_df = pd.read_excel(smb3_path, engine='openpyxl')
#gcc_df = pd.read_excel(gcc_path, engine='openpyxl')

# Clean up SalesProjectID_ContractNumber by removing decimals
for df_temp in [enterprise_df, smb1_df, smb2_df, smb3_df]:
    df_temp["SalesProjectID_ContractNumber"] = (
        df_temp["SalesProjectID_ContractNumber"]
        .astype(str)
        .str.replace(r"\.0+$", "", regex=True)  # Remove trailing .0
        .str.strip()
    )

print("📊 Rows in enterprise_df:", len(enterprise_df))
print("📊 Rows in smb1_df:", len(smb1_df))
print("📊 Rows in smb2_df:", len(smb2_df))
print("📊 Rows in smb3_df:", len(smb3_df))
print("🧮 Total before merge:", len(enterprise_df) + len(smb1_df) + len(smb2_df)+ len(smb3_df))

# === Combine all into one full newmerged dataset with all original columns ===
newmerged = pd.concat([enterprise_df, smb1_df, smb2_df, smb3_df], ignore_index=True)

merge_keys = ["ServiceID_Crid", "SalesProjectID_ContractNumber"]
newmerged = newmerged.drop_duplicates(subset=merge_keys)
print("🔄 Rows in newmerged (after concat + dedup):", len(newmerged))

# === Select just the enrichment columns from df ===
columns_to_add = merge_keys + [
    "Date signed",
    "Planning complete date",
    "Design complete date",
    "Correct RFS",
    "Actual RFS",
    "Expected RFS",
    "Deliveryconfirmationdate_Configsignoffdate",
    "Customer Agreed RFS",
    "CUSTOMERCONFIRMEDDELIVERYDATE",
    "CUSTOMERDEFERREDDELIVERYDATE",
    "ReadyForBilling",
    "ProjectTypeValue",
    "SOURCE"
]
df_subset = df[columns_to_add].copy()

# === Normalize merge keys: strip spaces and convert to string ===
for col in merge_keys:
    df_subset[col] = df_subset[col].astype(str).str.strip()
    newmerged[col] = newmerged[col].astype(str).str.strip()

# === Check for duplicates in df_subset ===
dupes = df_subset.duplicated(subset=merge_keys, keep=False)
duplicate_rows = df_subset[dupes]

if not duplicate_rows.empty:
    print(f"🚨 Warning: {len(duplicate_rows)} duplicate rows found in df_subset based on merge keys!")
    duplicate_output_path = rf"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\DK backlog python\1. Produce backlog\{today_str}_Output_duplicates_found.xlsx"
    duplicate_rows.to_excel(duplicate_output_path, index=False)
    df_subset = df_subset.drop_duplicates(subset=merge_keys)
    print(f"✅ Duplicates dropped — remaining df_subset rows: {len(df_subset)}")
else:
    print("✅ No duplicates in df_subset based on merge keys.")

# === Merge enrichment from df_subset into newmerged ===
final = newmerged.merge(
    df_subset,
    on=merge_keys,
    how="left"
)

print ("satan")
print("✅ Rows in final after merging with df:", len(final))

# ✅ Filter for desired Project Stage values before exporting. LOOK AT THIS
final = final[final["Project Stage"].isin(["In Process", "Created", "RFS"])]

# === Export the final merged file ===
output_path = rf"C:\Users\Tony Yang\OneDrive - McKinsey & Company\Documents\Python\2. Output\{today_str}_1. Final__latest_backlog_output (remember to update the first columns).xlsx"
final.to_excel(output_path, index=False)

print("✅ Final file exported successfully.")

# === Modify the final file: clean cols A-C, insert new cols A-B, and update A1–E1 ===
wb = load_workbook(output_path)
ws = wb.active

# Step 1: Clear contents in columns A, B, C (rows 2 onward)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.value = None

# Step 2: Insert two empty columns at the beginning
ws.insert_cols(1, 2)  # This shifts all columns to the right by 2

# Step 3: Set headers in cells A1 through E1
ws["A1"] = "Latest backlog update:"
ws["C1"] = "Input: Which delivery step is the order in? (Drop down list)"
ws["D1"] = "Input: What is the status of the order? (Drop down list)"
ws["E1"] = "(Optional) Input: Comment"

# Step 4: Save updated Excel file
wb.save(output_path)

print("✅ Columns A–E updated with custom values and cleaned.")